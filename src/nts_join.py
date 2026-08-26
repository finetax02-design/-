"""국세청 전자세금계산서 원본과 위하고 분개결과를 조인해 학습 데이터를 만든다.

왜 필요한가:
    위하고 화면 엑셀에는 위하고 내부 거래처코드(00411 등)만 있고 사업자등록번호가 없다.
    실제 운영에서 들어오는 입력은 국세청에서 내려받은 원본 파일이고, 거기엔
    사업자등록번호는 있어도 위하고 거래처코드가 없다.
    두 파일을 붙여야 '사업자등록번호 -> 계정과목' 룰을 만들 수 있다.

조인 키:
    (작성일자 MM-DD, 공급가액, 세액)
    승인번호가 위하고 엑셀에 없어서 금액+일자로 붙인다. 실측 조인율 100%.

사용법:
    python nts_join.py --nts 국세청원본.xls --wehago 위하고내역.xlsx --out dataset.csv
"""
from __future__ import annotations

import argparse
import collections
import csv
import warnings
from pathlib import Path

import openpyxl
import xlrd

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 국세청 파일은 '상호'/'대표자명'/'주소'/'종사업장번호'가 공급자·공급받는자 두 번 나온다.
# 헤더 이름으로 dict 를 만들면 뒤엣것이 앞엣것을 덮어쓰므로 열 위치로 직접 읽는다.
NTS_HEADER_ROW = 5
NTS_COL = {
    "작성일자": 0,
    "승인번호": 1,
    "공급자사업자번호": 4,
    "공급자상호": 6,
    "공급자주소": 8,
    "공급가액": 15,
    "세액": 16,
    "품목명": 26,
}

WEHAGO_COL = {
    "date": "일자",
    "code": "Code",
    "vendor": "거래처",
    "vat_type": "유형",
    "supply": "공급가액",
    "vat": "부가세",
    "debit": "차변계정",
    "credit": "대변계정",
    "status": "전표상태",
}
UNPOSTED = {"미추천", "확정가능", "", None}


def read_nts(path: Path) -> list[dict]:
    """국세청 전자세금계산서 목록조회 파일(.xls)을 읽는다."""
    sheet = xlrd.open_workbook(path).sheet_by_name("세금계산서")
    out = []
    for r in range(NTS_HEADER_ROW + 1, sheet.nrows):
        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if not str(row[0]).strip():
            continue
        get = lambda k: row[NTS_COL[k]]  # noqa: E731
        out.append({
            "작성일자": str(get("작성일자")),
            "승인번호": str(get("승인번호")).strip(),
            "사업자번호": str(get("공급자사업자번호")).strip(),
            "공급자상호": str(get("공급자상호")).strip(),
            "공급자주소": str(get("공급자주소")).strip(),
            "공급가액": int(round(float(get("공급가액") or 0))),
            "세액": int(round(float(get("세액") or 0))),
            "품목명": str(get("품목명")).strip(),
        })
    return out


def read_wehago(path: Path) -> list[dict]:
    """위하고 전자세금계산서 화면의 엑셀 내려받기 파일을 읽는다."""
    ws = openpyxl.load_workbook(path, data_only=True).active
    header = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(header, row))
        if not rec.get(WEHAGO_COL["code"]):
            continue  # 소계/합계 행
        out.append(rec)
    return out


def join(nts: list[dict], wehago: list[dict]) -> tuple[list[dict], list[dict]]:
    """(일자, 공급가액, 세액) 으로 두 데이터를 붙인다.

    같은 키가 여러 건일 수 있어(동일 거래처 같은 금액 반복) 큐에서 하나씩 꺼내 쓴다.
    """
    index = collections.defaultdict(collections.deque)
    for w in wehago:
        key = (str(w[WEHAGO_COL["date"]]),
               int(w[WEHAGO_COL["supply"]] or 0),
               int(w[WEHAGO_COL["vat"]] or 0))
        index[key].append(w)

    matched, unmatched = [], []
    for n in nts:
        key = (n["작성일자"][5:], n["공급가액"], n["세액"])
        queue = index.get(key)
        if not queue:
            unmatched.append(n)
            continue
        w = queue.popleft()
        matched.append({
            **n,
            "위하고거래처코드": w[WEHAGO_COL["code"]],
            "위하고거래처명": w[WEHAGO_COL["vendor"]],
            "계정과목": w[WEHAGO_COL["debit"]],
            "대변계정": w[WEHAGO_COL["credit"]],
            "공제구분": w[WEHAGO_COL["vat_type"]],
            "전표상태": w[WEHAGO_COL["status"]],
        })
    return matched, unmatched


def evaluate(rows: list[dict], target: str) -> dict:
    """사업자번호별 최빈값으로 target 을 예측했을 때의 정확도를 leave-one-out 으로 잰다."""
    by_biz = collections.defaultdict(list)
    for r in rows:
        by_biz[r["사업자번호"]].append(r)

    hit = miss = new = 0
    wrong = []
    for group in by_biz.values():
        for i, rec in enumerate(group):
            others = [x[target] for j, x in enumerate(group) if j != i]
            if not others:
                new += 1
                continue
            pred = collections.Counter(others).most_common(1)[0][0]
            if pred == rec[target]:
                hit += 1
            else:
                miss += 1
                wrong.append((rec, pred))
    return {"hit": hit, "miss": miss, "new": new, "wrong": wrong, "vendors": len(by_biz)}


def main() -> None:
    ap = argparse.ArgumentParser(description="국세청 원본 + 위하고 분개결과 조인")
    ap.add_argument("--nts", required=True, type=Path, help="국세청 전자세금계산서 원본 .xls")
    ap.add_argument("--wehago", required=True, type=Path, help="위하고 화면 내려받기 .xlsx")
    ap.add_argument("--out", type=Path, default=Path("dataset.csv"))
    args = ap.parse_args()

    nts = read_nts(args.nts)
    wehago = read_wehago(args.wehago)
    matched, unmatched = join(nts, wehago)

    print("=" * 74)
    print(f"국세청 {len(nts)}건  ↔  위하고 {len(wehago)}건")
    print(f"  조인 성공 {len(matched)}건 ({len(matched) / len(nts) * 100:.1f}%)  실패 {len(unmatched)}건")
    print("=" * 74)
    for n in unmatched[:10]:
        print(f"  미조인: {n['작성일자']} {n['공급자상호'][:20]:<22} {n['공급가액']:>12,}")

    posted = [r for r in matched if r["계정과목"] not in UNPOSTED]
    print(f"\n분개 완료된 학습 대상: {len(posted)}건")

    for target, label in (("계정과목", "사업자번호 → 계정과목"), ("공제구분", "사업자번호 → 공제구분")):
        res = evaluate(posted, target)
        total = res["hit"] + res["miss"] + res["new"]
        print(f"\n[{label}]  거래처 {res['vendors']}곳")
        print(f"    적중   {res['hit']:5d}건 ({res['hit'] / total * 100:5.1f}%)")
        print(f"    오답   {res['miss']:5d}건 ({res['miss'] / total * 100:5.1f}%)")
        print(f"    신규   {res['new']:5d}건 ({res['new'] / total * 100:5.1f}%)  ← LLM 담당 구간")
        for rec, pred in res["wrong"][:12]:
            print(f"      {rec['작성일자'][5:]} {rec['공급자상호'][:16]:<18} "
                  f"{rec['품목명'][:30]:<32} 예측={pred} 실제={rec[target]}")

    fields = ["작성일자", "승인번호", "사업자번호", "공급자상호", "공급자주소", "품목명",
              "공급가액", "세액", "위하고거래처코드", "위하고거래처명",
              "계정과목", "대변계정", "공제구분", "전표상태"]
    with args.out.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(matched)
    print(f"\n학습 데이터 저장: {args.out} ({len(matched)}건)")


if __name__ == "__main__":
    main()
