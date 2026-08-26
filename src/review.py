"""검토가 필요한 건만 골라 엑셀로 뽑는다.

원리:
    위하고의 자동매칭과 이 프로그램의 룰은 서로 독립적으로 계정과목을 판단한다.
    둘이 같은 답을 내면 굳이 사람이 볼 이유가 없고, 다르면 그 건만 보면 된다.
    위하고를 조작하지 않는다. 읽기만 하고, 결과는 별도 엑셀로 나온다.

검토 사유:
    계정과목 불일치 - 위하고가 채운 계정과 룰의 예측이 다르다
    공제구분 불일치 - 과세/불공 판정이 다르다
    이력 없음      - 신규 거래처이고 품명도 처음이라 룰이 판단을 못 한다
    미추천         - 위하고가 아예 계정을 못 채웠다

사용법:
    python review.py --nts 국세청.xls --wehago 위하고.xlsx --history 과거.csv --out 검토목록.xlsx
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nts_join import UNPOSTED, join, read_nts, read_wehago
from predict import Model

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
REASON_FILL = {
    "미추천": PatternFill("solid", fgColor="FFC7CE"),
    "계정과목 불일치": PatternFill("solid", fgColor="FFE699"),
    "공제구분 불일치": PatternFill("solid", fgColor="FFE699"),
    "이력 없음(신규)": PatternFill("solid", fgColor="DDEBF7"),
}
COLUMNS = ["일자", "거래처", "품목명", "공급가액",
           "위하고 계정과목", "룰 예측 계정과목",
           "위하고 공제", "룰 예측 공제", "검토 사유"]
WIDTHS = [10, 24, 40, 14, 18, 18, 10, 10, 18]


def classify(row: dict, acct_pred: str | None, vat_pred: str | None) -> str | None:
    """이 건을 사람이 봐야 하는지, 본다면 왜인지 판정한다."""
    if row["계정과목"] in UNPOSTED:
        return "미추천"
    if acct_pred is None:
        return "이력 없음(신규)"
    if acct_pred != row["계정과목"]:
        return "계정과목 불일치"
    if vat_pred is not None and vat_pred != row["공제구분"]:
        return "공제구분 불일치"
    return None


def write_xlsx(flagged: list[tuple[dict, str | None, str | None, str]],
               total: int, out: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검토목록"

    ws.append([f"검토 대상 {len(flagged)}건 / 전체 {total}건 "
               f"({len(flagged) / total * 100:.1f}%)  — 나머지는 위하고와 룰이 일치"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(COLUMNS)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row, acct, vat, reason in flagged:
        ws.append([row["작성일자"][5:], row["공급자상호"], row["품목명"], row["공급가액"],
                   row["계정과목"], acct or "-", row["공제구분"], vat or "-", reason])
        for cell in ws[ws.max_row]:
            cell.fill = REASON_FILL[reason]
        ws.cell(ws.max_row, 4).number_format = "#,##0"

    for i, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{ws.max_row}"
    wb.save(out)


def main() -> None:
    ap = argparse.ArgumentParser(description="검토가 필요한 건만 골라 엑셀로 출력")
    ap.add_argument("--nts", required=True, type=Path, help="국세청 원본 .xls")
    ap.add_argument("--wehago", required=True, type=Path, help="위하고 내려받기 .xlsx")
    ap.add_argument("--history", type=Path, help="과거 학습 데이터 .csv (없으면 대상 자체로 학습)")
    ap.add_argument("--out", type=Path, default=Path("검토목록.xlsx"))
    ap.add_argument("--month", help="이번에 검토할 월만 걸러낸다 (예: 06)")
    args = ap.parse_args()

    rows, _ = join(read_nts(args.nts), read_wehago(args.wehago))
    if args.month:
        rows = [r for r in rows if r["작성일자"][5:7] == args.month]
    if args.history:
        train = [r for r in csv.DictReader(args.history.open(encoding="utf-8-sig"))
                 if r["계정과목"] not in UNPOSTED]
    else:
        train = [r for r in rows if r["계정과목"] not in UNPOSTED]

    acct_model = Model("계정과목").fit(train)
    vat_model = Model("공제구분").fit(train)

    flagged = []
    for row in rows:
        acct, _, _ = acct_model.predict(row)
        vat, _, _ = vat_model.predict(row)
        reason = classify(row, acct, vat)
        if reason:
            flagged.append((row, acct, vat, reason))

    write_xlsx(flagged, len(rows), args.out)
    print(f"전체 {len(rows)}건 중 검토 대상 {len(flagged)}건 ({len(flagged) / len(rows) * 100:.1f}%)")
    for reason in ("미추천", "계정과목 불일치", "공제구분 불일치", "이력 없음(신규)"):
        n = sum(1 for f in flagged if f[3] == reason)
        if n:
            print(f"    {reason:<16} {n:4d}건")
    print(f"저장: {args.out}")


if __name__ == "__main__":
    main()
