"""위하고 전자세금계산서 내역을 분석해 자동분개 룰의 커버리지를 계산한다.

입력: 위하고 T > 전자세금계산서 화면에서 엑셀로 내려받은 파일
      (컬럼: 일자 Code 거래처 유형 품명 공급가액 부가세 합계 차변계정 대변계정 관리 전표상태)

출력: 콘솔 리포트 + rules.json (거래처코드 -> 계정과목/공제구분 룰 테이블)

사용법:
    python analyze.py <엑셀파일> [엑셀파일 ...] [--out rules.json]
"""
from __future__ import annotations

import argparse
import collections
import json
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 위하고 엑셀의 컬럼명. 프로그램 버전이 바뀌어 컬럼명이 달라지면 여기만 고치면 된다.
COL = {
    "date": "일자",
    "code": "Code",
    "vendor": "거래처",
    "vat_type": "유형",       # 과세 / 불공 / 면세 ...
    "item": "품명",
    "supply": "공급가액",
    "vat": "부가세",
    "total": "합계",
    "debit": "차변계정",
    "credit": "대변계정",
    "status": "전표상태",
}

# 분개가 아직 안 붙은 행에서 차변계정 자리에 들어오는 값들
UNPOSTED = {"미추천", "확정가능", "", None}


def load(paths: list[Path]) -> list[dict]:
    """엑셀 파일들을 읽어 dict 리스트로 합친다."""
    out = []
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        missing = [k for k in (COL["code"], COL["debit"]) if k not in header]
        if missing:
            sys.exit(f"[오류] {path.name}: 필수 컬럼 없음 {missing}\n  실제 헤더: {header}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(header, row))
            if not rec.get(COL["code"]):
                continue  # '합계(거래처:15)' 같은 소계/합계 행에는 거래처코드가 없다
            out.append(rec)
    return out


def month_of(rec: dict) -> str:
    """일자에서 월(MM)만 뽑는다. '01-02' 문자열과 datetime 둘 다 처리."""
    v = rec.get(COL["date"])
    if hasattr(v, "month"):
        return f"{v.month:02d}"
    return str(v)[:2] if v else "??"


def split_posted(records: list[dict]) -> tuple[list[dict], list[dict]]:
    """분개 완료분(정답 데이터)과 미분개분으로 나눈다."""
    posted, unposted = [], []
    for r in records:
        (unposted if r.get(COL["debit"]) in UNPOSTED else posted).append(r)
    return posted, unposted


def build_rules(posted: list[dict]) -> dict:
    """거래처코드별로 계정과목/공제구분 룰을 만든다.

    같은 거래처가 항상 같은 계정과목을 쓰면 그 계정을 확정 룰로,
    갈리면 최빈값과 함께 '검토 필요'로 표시한다.
    """
    by_code = collections.defaultdict(list)
    for r in posted:
        by_code[r[COL["code"]]].append(r)

    rules = {}
    for code, rows in by_code.items():
        debit = collections.Counter(r[COL["debit"]] for r in rows)
        credit = collections.Counter(r[COL["credit"]] for r in rows)
        vat = collections.Counter(r[COL["vat_type"]] for r in rows)
        top_acct, top_n = debit.most_common(1)[0]
        rules[str(code)] = {
            "거래처": rows[0][COL["vendor"]],
            "건수": len(rows),
            "계정과목": top_acct,
            "계정과목_확신도": round(top_n / len(rows), 3),
            "계정과목_분포": dict(debit),
            "대변계정": credit.most_common(1)[0][0],
            "공제구분_분포": dict(vat),
            "단일계정": len(debit) == 1,
        }
    return rules


def leave_one_out(posted: list[dict]) -> dict:
    """각 건을 하나씩 빼고, 같은 거래처의 나머지 이력으로 계정과목을 맞춰본다.

    '신규 거래처가 들어왔을 때 룰이 얼마나 버티는가'를 실측하는 방식.
    """
    by_code = collections.defaultdict(list)
    for r in posted:
        by_code[r[COL["code"]]].append(r)

    hit = miss = 0
    new_vendor = []
    wrong = []
    for rows in by_code.values():
        for i, rec in enumerate(rows):
            others = [x[COL["debit"]] for j, x in enumerate(rows) if j != i]
            if not others:
                new_vendor.append(rec)
                continue
            pred = collections.Counter(others).most_common(1)[0][0]
            if pred == rec[COL["debit"]]:
                hit += 1
            else:
                miss += 1
                wrong.append((rec, pred))
    return {"hit": hit, "miss": miss, "new_vendor": new_vendor, "wrong": wrong}


def report(records: list[dict]) -> dict:
    posted, unposted = split_posted(records)
    print("=" * 74)
    print(f"전체 {len(records)}건  |  분개완료 {len(posted)}건  |  미분개 {len(unposted)}건")
    print("=" * 74)

    print("\n[1] 계정과목 분포")
    for acct, n in collections.Counter(r[COL["debit"]] for r in posted).most_common():
        print(f"    {n:5d}건  {acct}")

    print("\n[2] 공제구분(유형) — 월별")
    by_month = collections.defaultdict(collections.Counter)
    for r in posted + unposted:
        by_month[month_of(r)][r[COL["vat_type"]]] += 1
    for mo in sorted(by_month):
        c = by_month[mo]
        detail = "  ".join(f"{k}={v}" for k, v in c.most_common() if k)
        print(f"    {mo}월  {detail}")

    rules = build_rules(posted)
    single = [v for v in rules.values() if v["단일계정"]]
    multi = [v for v in rules.values() if not v["단일계정"]]
    print(f"\n[3] 거래처 {len(rules)}곳")
    print(f"    단일계정 거래처 : {len(single):4d}곳 / {sum(v['건수'] for v in single):5d}건")
    print(f"    다계정 거래처   : {len(multi):4d}곳 / {sum(v['건수'] for v in multi):5d}건")
    for v in sorted(multi, key=lambda x: -x["건수"]):
        dist = ", ".join(f"{a}×{n}" for a, n in v["계정과목_분포"].items())
        print(f"      · {str(v['거래처'])[:26]:<28} {v['건수']:3d}건 → {dist}")

    loo = leave_one_out(posted)
    total = loo["hit"] + loo["miss"] + len(loo["new_vendor"])
    print(f"\n[4] L1 룰(거래처코드 → 계정과목) 실측  n={total}")
    if total:
        print(f"    적중        {loo['hit']:5d}건  ({loo['hit'] / total * 100:5.1f}%)")
        print(f"    오답        {loo['miss']:5d}건  ({loo['miss'] / total * 100:5.1f}%)")
        print(f"    이력없음    {len(loo['new_vendor']):5d}건  ({len(loo['new_vendor']) / total * 100:5.1f}%)  ← LLM이 담당할 구간")
    for rec, pred in loo["wrong"][:10]:
        print(f"      오답: {str(rec[COL['vendor']])[:22]:<24} 예측={pred} 실제={rec[COL['debit']]}")

    print("\n[5] 이력 없는 거래처 (신규 — 품명으로 추론해야 하는 건)")
    for rec in loo["new_vendor"][:30]:
        print(f"    {str(rec[COL['vendor']])[:22]:<24} | {str(rec[COL['item']])[:32]:<34} → {rec[COL['debit']]}")

    if unposted:
        print(f"\n[6] 미분개 {len(unposted)}건 — 기존 룰로 커버되는가")
        covered = [r for r in unposted if str(r[COL["code"]]) in rules]
        print(f"    룰 있음 : {len(covered):4d}건 ({len(covered) / len(unposted) * 100:.1f}%) → 즉시 자동분개 가능")
        print(f"    룰 없음 : {len(unposted) - len(covered):4d}건 → LLM 판단 필요")
        for r in unposted:
            if str(r[COL["code"]]) not in rules:
                print(f"      신규: {str(r[COL['vendor']])[:22]:<24} | {str(r[COL['item']])[:34]}")
    return rules


def main() -> None:
    ap = argparse.ArgumentParser(description="위하고 전자세금계산서 자동분개 룰 커버리지 분석")
    ap.add_argument("files", nargs="+", type=Path, help="위하고에서 내려받은 엑셀 파일")
    ap.add_argument("--out", type=Path, default=Path("rules.json"), help="룰 테이블 저장 경로")
    args = ap.parse_args()

    records = load(args.files)
    rules = report(records)
    args.out.write_text(json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n룰 테이블 저장: {args.out}  ({len(rules)}개 거래처)")


if __name__ == "__main__":
    main()
